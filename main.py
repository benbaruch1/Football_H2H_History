from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
import time

path_of_driver = r"C:\python\chromedriver.exe"
options = Options()
options.headless = False
driver = webdriver.Chrome(path_of_driver, options=options)

team1 = input("Please choose 1st team: ")
team2 = input("Please choose 2nd team: ")

driver.get("https://matchstat.com/football/head-to-head")
time.sleep(2)
close_popup = driver.find_element_by_xpath('//button[@class="close"]')
close_popup.click()

input_team1 = driver.find_element_by_xpath('//input[@id="h2h-search-player1"]')
input_team2 = driver.find_element_by_xpath('//input[@id="h2h-search-player2"]')
time.sleep(2)
input_team1.send_keys(team1)
time.sleep(2)
input_team1.send_keys(Keys.ARROW_DOWN)
input_team1.send_keys(Keys.RETURN)
time.sleep(2)
input_team2.send_keys(team2)
time.sleep(2)
input_team2.send_keys(Keys.ARROW_DOWN)
input_team2.send_keys(Keys.RETURN)
time.sleep(3)

try:
    get_table_h2h_row = driver.find_element_by_xpath('//table[@class="table fb-fixture-table"]')
except NoSuchElementException:
    print("Input error.")
    driver.quit()
    exit()

amount_of_games = len(get_table_h2h_row.text.split('\n'))

if amount_of_games <= 1:
    print("The teams never played against each other or no stats available.")
    driver.quit()
    exit()

temp = []
games_list = []
for i in range(2, amount_of_games + 1):
    for j in range(1, 6):
        get_each_info = driver.find_element_by_xpath('//table[@class="table fb-fixture-table"]/tbody/tr[' + str(i) + ']/td[' + str(j) + ']')
        temp.append(get_each_info.text)
    games_list.append(temp)
    temp = []

wb = Workbook()
ws = wb.active
ws['A1'] = "Date"
ws['C1'] = "Competition"
ws['E1'] = "Home"
ws['G1'] = "Result"
ws['I1'] = "Away"

i = 2
for game in games_list:
    if game[3] == '':  # If score is empty, continue. (Only games that already played)
        continue
    ws['A' + str(i) + ''] = game[0]
    ws['C' + str(i) + ''] = game[1]
    ws['E' + str(i) + ''] = game[2]
    ws['G' + str(i) + ''] = game[3]
    ws['I' + str(i) + ''] = game[4]
    i += 1

wb.save(r"C:\Users\Ben\Desktop\GameHistory.xlsx")

driver.quit()
